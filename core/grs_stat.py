import pandas as pd
import sqlite3
import openpyxl as oxl
import matplotlib.pyplot as plt
from tkinter import *
from tkinter import scrolledtext
from tkinter.ttk import Combobox
import random
import CoolProp.CoolProp as CP

from core.entity.Gas import Gas

design_pressure = 5.39
start_pressure = 1.5
pressure_steps = 100
pressure_step = (design_pressure - start_pressure) / pressure_steps
start_temperature = 2
end_temperature = 13
temperature_step = 1


component_list = ['Methane', 'Ethane', 'Propane', 'Isobutane', 'Butane',
                  'Isopentane', 'Pentane', 'Hexane', 'Oxygen',
                  'Nitrogen', 'CarbonDioxide']
component_list_for_query = ['methane', 'ethane', 'propane', 'isobutane', 'butane',
                            'isopentane', 'pentane', 'hexane', 'oxygen',
                            'nitrogen', 'carbon_dioxide']


def get_grs_name():
    connection = sqlite3.connect('grs_database.db')
    cur = connection.cursor()
    cur.execute("""SELECT DISTINCT name_grs FROM grs""")
    grs_list = [grs[0] for grs in cur.fetchall()]
    connection.close()
    return grs_list


def get_composition(name_grs, normolize=True):
    connection = sqlite3.connect('grs_database.db')
    cur = connection.cursor()
    component_list_query = ', '.join(component for component in component_list_for_query)
    cur.execute(f"""SELECT {component_list_query} FROM composition
                    WHERE grs_id = (SELECT grs_id FROM grs 
                    WHERE name_grs = '{name_grs}');""")
    values = cur.fetchone()
    if normolize == True:
        while abs(sum(values) - 1) > 10e-10:
            values = [component + ((1 - sum(values)) * component) for component in values]
    composition = dict(zip(component_list, values))
    connection.close()
    return composition


database_path = r'C:\Users\Идель\PycharmProjects\pythonProject2\grs_database.db'

months = ('01', '02', '03', '04',
          '05', '06', '07', '08',
          '09', '10', '11', '12')
seasons = {'Весь период': -1, 'Зима': 0, 'Весна': 1, 'Лето': 2, 'Осень': 3}


def drop_table(database_path):
    connection = sqlite3.connect(database_path)
    cur = connection.cursor()
    cur.execute(f"""DROP TABLE grs_stat""")
    connection.commit()
    connection.close()


def record_data():
    def export_to_sqlite():

        database_path = db_path_entry.get()
        if database_path == '':
            database_path = r'C:\Users\Идель\PycharmProjects\pythonProject2\grs_database.db'
            message = 'Прописан стандартный путь к базе данных\n'
        grs_name = grs_cmb.get()
        if grs_name == '':
            message = 'Вы не ввели название ГРС!\n'

        exl_file_name = grs_name + '_статистика.xlsx'
        connection = sqlite3.connect(database_path)
        cur = connection.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS grs_stat(
        data_id INTEGER PRIMARY KEY AUTOINCREMENT,
        grs_id INTEGER NOT NULL,
        name_output VARCHAR(15) NOT NULL,
        date TIMESTAMP NOT NULL,
        nominal_capacity DECIMAL(8, 4),
        day_capacity DECIMAL(12, 4) NOT NULL,
        hour_capacity DECIMAL(10, 4) NOT NULL,
        pressure DECIMAL(4, 4) NOT NULL,
        temperature DECIMAL(4, 4),
        actual_flow DECIMAL(10, 2),
        FOREIGN KEY (grs_id) REFERENCES grs(grs_id) ON DELETE CASCADE
        );""")

        try:
            file_to_read = oxl.load_workbook(exl_file_name, data_only=True)
        except FileNotFoundError:
            txtwin.insert(INSERT, f'По указанному пути файла {exl_file_name} не обнаружено!\n')
            return
        sheets = file_to_read.sheetnames

        print(sheets)

        for sheet in sheets:
            for row in range(4, file_to_read[sheet].max_row + 1):
                data = []
                for col in range(1, 7):
                    value = file_to_read[sheet].cell(row, col).value
                    if value is None:
                        break
                    else:
                        data.append(value)
                try:
                    cur.execute(f"""INSERT INTO grs_stat(name_output, date, nominal_capacity, 
                    day_capacity, hour_capacity, pressure, temperature, grs_id)
                    VALUES(?, ?, ?, ?, ?, ?, ?,
                    (SELECT grs_id FROM grs WHERE name_grs = '{grs_name}'));""",
                                (sheet, data[0], data[1], data[2], data[3], data[4], data[5]))
                except IndexError:
                    pass
                    # print('В исходной таблице Excel имеются пустые строчки')

            txtwin.insert(INSERT, f'База данных успешно обновлена! Обнаружено '
                                  f'{file_to_read[sheet].max_row} на вкладке {sheet}\n')

        """если ты надобавлял дубликатов, то этот запрос должен их тут же удалить"""
        cur.execute("""DELETE FROM grs_stat
        WHERE data_id not in (
        SELECT min(data_id) FROM grs_stat 
        GROUP BY name_output, date, grs_id);""")
        connection.commit()
        connection.close()

    def calc_actual_flow():
        connection = sqlite3.connect('grs_database.db')
        cur = connection.cursor()
        grs_name = grs_cmb.get()

        cur.execute(f"""SELECT hour_capacity, pressure, temperature, data_id FROM grs_stat 
        WHERE grs_id = (SELECT grs_id FROM grs WHERE name_grs = '{grs_name}')""")
        states = cur.fetchall()
        rates = [state[0] * 1000 for state in states]
        pressures = [state[1] * 98100 / 1000000 for state in states]
        temperatures = [state[2] for state in states]
        data_id_list = [id[3] for id in states]
        gas = Gas(get_composition(grs_cmb.get()), temperatures[0], pressures[0], rates[0]) #первичная инициализация
        actual_flows = []
        for i in range(len(states)):
            try:
                if rates[i] == 0:  # клёво я придумал оптимизировать эти расчёты, если расхода нет, зачем инициализировать весь газ
                    actual_flows.append(0)
                else:
                    actual_flows.append(gas.get_actual_rate(temperatures[i], pressures[i]))
            except ValueError:
                actual_flows.append(actual_flows[-1] * rates[i] / rates[i - 1] *
                                    (pressures[i - 1] + 0.101325) / (pressures[i] + 0.101325) *
                                    (temperatures[i] + 273.15) / (273.15 + temperatures[i - 1]))
            cur.execute(f"""UPDATE grs_stat SET actual_flow = {actual_flows[i]} WHERE data_id = {data_id_list[i]}""")
            connection.commit()
            print(f'Прогресс {round(100 * i / len(states), 2)}%')

    record_data_root = Tk()
    record_data_root.geometry('800x1000')
    record_data_root.title('Запись в базу данных')

    txtwin = scrolledtext.ScrolledText(record_data_root, width=80, height=15, bg="darkgreen", fg='white')
    txtwin.pack(pady=10)

    lbl = Label(record_data_root, text='Название ГРС')
    lbl.pack()

    grs_cmb = Combobox(record_data_root)
    grs_cmb['values'] = get_grs_name()
    grs_cmb.current(0)
    grs_cmb.pack()

    lbl = Label(record_data_root, text='Пусть к базе данных\nПуть по умолчанию вводить не обязательно')
    lbl.pack()
    db_path_entry = Entry(record_data_root)
    db_path_entry.pack()

    record_btn = Button(record_data_root, text='Запись', command=export_to_sqlite, bg='darkgreen', fg='white')
    record_btn.pack()

    calc_flow_btn = Button(record_data_root, text='Расчёт актуального расхода',
                           command=calc_actual_flow, bg='darkgreen', fg='white')
    calc_flow_btn.pack()

    record_data_root.mainloop()


# drop_table(database_path)

# export_to_sqlite(database_path, grs_name)


"""Главная функция, создающая интерфейс и содержащая в себе все основные функции"""

def make_stat_root():
    """когда изменяется название ГРС в комбобоксе, автоматически обновляются варианты выходов"""

    def on_combobox_select(event):
        connection = sqlite3.connect(database_path)
        grs_name = event.widget.get()
        """этот запрос ищет названия выходов, чтобы ты мог выбрать какой тебе вписать"""
        outputs_names = list(pd.read_sql(f"SELECT DISTINCT name_output as'Выход' "
                                         f"FROM grs_stat "
                                         f"WHERE grs_id = ("
                                         f"SELECT grs_id "
                                         f"FROM grs "
                                         f"WHERE name_grs = '{grs_name}');", connection)['Выход'])
        if len(outputs_names) > 1:
            outputs_names.append('Вход')
        combo_outputs['values'] = outputs_names
        combo_outputs.set(outputs_names[0])

    def get_stat_db():
        output_name = combo_outputs.get()
        name_grs = combo_grs.get()
        connection = sqlite3.connect(database_path)

        key = combo_season.get()
        if key == 'Весь период':
            if output_name == 'Вход':
                query = f"""SELECT date AS 'Дата', ROUND( pressure * 0.0981, 4) as '{parameters_list[0]}',
                 ROUND(SUM(hour_capacity)*1000, 0) AS '{parameters_list[1]}', temperature AS '{parameters_list[2]}', 
                 ROUND(SUM(actual_flow), 0) AS '{parameters_list[3]}' FROM grs_stat
                                WHERE grs_id  = (SELECT grs_id FROM grs WHERE name_grs = '{name_grs}')
                                GROUP BY date;"""
            else:
                query = f"""SELECT date as 'Дата',
                            ROUND( pressure * 0.0981, 4) as '{parameters_list[0]}',
                              ROUND(hour_capacity * 1000, 1) as '{parameters_list[1]}',
                              temperature AS '{parameters_list[2]}', actual_flow AS '{parameters_list[3]}'
                               FROM grs_stat WHERE name_output = '{output_name}';"""
        else:
            i = seasons[key]
            if output_name == 'Вход':
                query = f"""SELECT date AS 'Дата', ROUND( pressure * 0.0981, 4) as '{parameters_list[0]}',
                 ROUND(SUM(hour_capacity)*1000, 0) AS '{parameters_list[1]}', temperature AS '{parameters_list[2]}',
                  ROUND(SUM(actual_flow), 0) AS '{parameters_list[3]}'FROM grs_stat
                        WHERE (date LIKE '____-{months[3 * i - 1]}-__%' 
                        OR  date LIKE'____-{months[3 * i]}-__%' 
                        OR date LIKE '____-{months[3 * i + 1]}-__%') 
                        AND (grs_id = (SELECT grs_id FROM grs WHERE name_grs = '{name_grs}'))
                        GROUP BY date;
                        """
            else:
                query = f"""SELECT date AS 'Дата', ROUND( pressure * 0.0981, 4) as '{parameters_list[0]}',
                 ROUND(SUM(hour_capacity)*1000, 0) AS '{parameters_list[1]}', temperature AS '{parameters_list[2]}', 
                   actual_flow AS '{parameters_list[3]}' FROM grs_stat
                        WHERE (date LIKE '____-{months[3 * i - 1]}-__%' 
                        OR  date LIKE'____-{months[3 * i]}-__%' 
                        OR date LIKE '____-{months[3 * i + 1]}-__%') 
                        AND (name_output = '{output_name}')
                        GROUP BY date;
                        """

        return pd.read_sql(query, connection)

    def filter(df):
        """эта функция проводит фильтр значений (первый аргумент) по вычисляемому тут порогу (второй аргумент)"""
        column = combo_column.get()
        std = df[column].std()
        threshold = float(format(threshold_entry.get().replace(',', '.')))
        limit = threshold * std
        return df[abs(df[column] - df[column].mean()) <= limit]

    def scam(df):
        """эта функция проводит фильтр значений (первый аргумент) по вычисляемому тут порогу (второй аргумент)"""
        column = combo_column.get()
        std = df[column].std()
        threshold = float(format(threshold_entry.get().replace(',', '.')))
        limit = threshold * std
        return df[abs(df[column] - df[column].mean()) > limit]

    def filtered_plot():
        # Построение графика
        name_grs = combo_grs.get()
        column = combo_column.get()
        season = combo_season.get()
        df = get_stat_db()
        filtered = filter(df).sort_values(by=['Дата'])
        scamed = scam(df).sort_values(by=['Дата'])
        threshold = float(format(threshold_entry.get().replace(',', '.')))
        limit = df[column].std() * threshold
        output_name = combo_outputs.get()
        plt.figure(figsize=(16, 8))
        plt.plot(filtered.index,
                 filtered[column],
                 label='Принятые значения',
                 marker='o',
                 linestyle='-',
                 color='blue',
                 markersize=3)
        plt.plot(scamed.index,
                 scamed[column],
                 marker='o',
                 label='Вылетевшие точки',
                 color='red',
                 linestyle='',
                 markersize=3)
        plt.axhline(y=df[column].mean(), color='gray', linestyle='--', label='Среднее значение')
        plt.axhline(y=df[column].mean() + limit, color='green', linestyle='--', label='Верхний порог')
        if column != 'Температура':
            if df[column].mean() - limit < 0:
                plt.axhline(y=0, color='green', linestyle='--', label='Нижний порог')
            else:
                plt.axhline(y=df[column].mean() - limit, color='green', linestyle='--', label='Нижний порог')
        plt.xlim(0, df.shape[0])
        plt.legend()
        plt.xlabel('Дата')
        mesuares = {parameters_list[0]: ', МПа (изб.)', parameters_list[1]: ', ст. м3/ч',
                    parameters_list[2]: ', °С', parameters_list[3]: ', м3/ч'}
        mesuare = mesuares[column]
        plt.ylabel(column + mesuare)
        plt.title(f'{column} газа {name_grs}, {output_name}, {season}')
        plt.grid(False)
        plt.show()

    def draw_plot():
        """считываем базу данных"""

        output_name = combo_outputs.get()
        column = combo_column.get()
        grs_name = combo_grs.get()
        season = combo_season.get()
        df = get_stat_db()

        df.sort_values(by=['Дата'])
        dates = pd.date_range(start=df['Дата'].min(), end=df['Дата'].max(), freq='D')
        if (len(dates)) > (len(list(df[column]))):
            dates = df.index
            day_mesuare = f'Дни сезона {season}'
        else:
            day_mesuare = 'Дата'
        plt.figure(figsize=(16, 8))
        plt.plot(dates, df[column], marker='o', linestyle='-', color='blue', markersize=3)
        plt.xlabel(day_mesuare)
        mesuares = {parameters_list[0]: ', МПа (изб.)', parameters_list[1]: ', ст. м3/ч',
                    parameters_list[2]: ', °С', parameters_list[3]: ', м3/ч'}
        mesuare = mesuares[column]
        plt.ylabel(column + mesuare)
        plt.title(f'{column} ГРС {grs_name}, {output_name}, {season}')
        plt.grid(linewidth=1, linestyle='-', color='grey')
        plt.show()

    def clean():
        txtwin.delete(1.0, END)

    def statistic():
        clean()
        grs_name = combo_grs.get()
        column = combo_column.get()
        output_name = combo_outputs.get()
        df = get_stat_db()

        average_min = round(df[column].min(), 4)
        average_mean = round(df[column].mean(), 4)
        average_max = round(df[column].max(), 4)
        if column == 'Давление':
            mesuare = 'МПа (изб.)'
        elif column == 'Расход':
            mesuare = 'ст. м3/ч'
        else:
            mesuare = ''
        txtwin.insert(INSERT, (f'Абсолютные значения за период {combo_season.get()}\n'
                               f'ГРС {grs_name}, параметр {column}:\n'
                               f'максимум {average_max} {mesuare}\n'
                               f'среднее {average_mean} {mesuare}\n'
                               f'минимум {average_min} {mesuare}\n\n'))
        filtered = filter(df)
        scamed = scam(df)
        average_min = round(filtered[column].min(), 4)
        average_mean = round(filtered[column].mean(), 4)
        average_max = round(filtered[column].max(), 4)
        txtwin.insert(INSERT, (f'Обработанные значения за период {combo_season.get()}\n'
                               f'ГРС {grs_name}, параметр {column}:\n'
                               f'максимум {average_max} {mesuare}\n'
                               f'среднее {average_mean} {mesuare}\n'
                               f'минимум {average_min} {mesuare}\n'
                               f'отброшенных точек - {scamed.shape[0]} ({round(scamed.shape[0] / df.shape[0] * 100, 2)}%)\n\n'))
        filtered_plot()

    connection = sqlite3.connect(database_path)

    cur = connection.cursor()

    try:
        cur.execute("""SELECT DISTINCT name_grs FROM grs WHERE grs_id IN (SELECT DISTINCT grs_id FROM grs_stat);""")
    except sqlite3.OperationalError:
        record_data()
        pass

    grs_combo = [grs[0] for grs in cur.fetchall()]  # получили все ГРС, которые были в базе данных по статистике ГРС
    print(grs_combo)

    if grs_combo == []:
        record_data()
    else:

        stat_root = Tk()
        stat_root.geometry('800x1000')
        stat_root.title('Статистика ГРС')

        """создали комбобокс с именами ГРС, по умолчанию там стоит первая ГРС в списке"""
        lbl = Label(stat_root, text='ГРС')
        lbl.pack()
        combo_grs = Combobox(stat_root, justify='center')
        combo_grs['values'] = grs_combo
        combo_grs.current(0)
        combo_grs.pack()
        combo_grs.bind('<<ComboboxSelected>>', on_combobox_select)

        """создали комбобокс с параметрами, по которым можно делать статистику"""
        lbl = Label(stat_root, text='Параметр')
        lbl.pack()
        combo_column = Combobox(stat_root, justify='center')
        parameters_list = ['Давление', 'Расход', 'Температура', 'Расход при р.у.']
        combo_column['values'] = parameters_list
        combo_column.current(0)
        combo_column.pack(pady=10)

        grs_name = combo_grs.get()
        outputs_names = list(pd.read_sql(f"SELECT DISTINCT name_output as'Выход' "
                                         f"FROM grs_stat "
                                         f"WHERE grs_id = ("
                                         f"SELECT grs_id "
                                         f"FROM grs "
                                         f"WHERE name_grs = '{grs_name}');", connection)['Выход'])
        if len(outputs_names) > 1:
            outputs_names.append('Вход')
        """создали комбобокс с названиями выходов ГРС, указанной в комбобоксе имен ГРС"""
        lbl = Label(stat_root, text='Выход')
        lbl.pack()
        combo_outputs = Combobox(stat_root, justify='center')
        combo_outputs['values'] = outputs_names
        combo_outputs.current(0)
        combo_outputs.pack(pady=10)

        txtwin = scrolledtext.ScrolledText(stat_root, width=80, height=25, bg="darkgreen", fg='white')
        txtwin.pack(pady=10)

        command_list = [draw_plot, statistic, clean, record_data]
        btn_list = ['График', 'Стат. Анализ', 'Очистить', 'Добавить новые данные']
        color_list = ['red', 'green', 'purple', 'blue', 'yellow', 'white']

        for btn in range(len(btn_list)):
            btn = Button(stat_root, text=btn_list[btn], command=command_list[btn],
                         bg='black', fg=random.choice(color_list))
            btn.pack(pady=10)
        lbl = Label(stat_root, text='Период анализа')
        lbl.pack()
        combo_season = Combobox(stat_root, justify='center')
        combo_season['values'] = list(seasons.keys())
        combo_season.current(0)
        combo_season.pack()

        lbl = Label(stat_root, text='количество доверительных стандартных отклонений')
        lbl.pack(pady=10)

        threshold_entry = Entry(stat_root, justify='center', relief='sunken')
        threshold_entry.pack()
        threshold_entry.insert(END, '3')

        stat_root.mainloop()


"""создает таблицу из excel файла с помощью панды"""
def export_to_sqlite_pandas(exl_file_name, database_path):
    df = pd.read_excel(exl_file_name)
    con = sqlite3.connect(database_path)
    df.to_sql('grs_stat', con, index=True, if_exists='replace')
    con.commit()
    con.close()
